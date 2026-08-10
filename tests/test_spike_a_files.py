"""Unit tests for the Spike A file generator.

The network is mocked with httpx's MockTransport — nothing here touches a real
host (CLAUDE.md 5.5). Failure paths matter more than the happy one: this
script's whole job is to refuse to produce a misleading test file.
"""

from __future__ import annotations

import csv
from collections.abc import Callable
from pathlib import Path

import httpx
import pytest
from openpyxl import load_workbook
from spikes.make_spike_a_files import (
    DEFAULT_PHOTO_URLS,
    HEADERS,
    MAX_IMAGE_URL_CHARS,
    ROW_TEXT,
    build_rows,
    verify_photo_url,
    write_csv,
    write_xlsx,
)

Handler = Callable[[httpx.Request], httpx.Response]


def _client(handler: Handler) -> httpx.Client:
    return httpx.Client(transport=httpx.MockTransport(handler))


def _ok(_request: httpx.Request) -> httpx.Response:
    return httpx.Response(200, headers={"content-type": "image/jpeg"}, content=b"\xff\xd8\xff")


def test_build_rows_pairs_text_with_urls() -> None:
    rows = build_rows(DEFAULT_PHOTO_URLS)
    assert len(rows) == len(ROW_TEXT)
    assert all(len(row) == len(HEADERS) for row in rows)
    assert rows[0][-1] == DEFAULT_PHOTO_URLS[0]


def test_build_rows_rejects_wrong_url_count() -> None:
    with pytest.raises(ValueError, match="exactly 2 photo URLs"):
        build_rows(("https://example.com/only-one.jpg",))


def test_verify_accepts_an_image() -> None:
    with _client(_ok) as client:
        assert verify_photo_url("https://example.com/a.jpg", client) is None


def test_verify_rejects_non_https() -> None:
    with _client(_ok) as client:
        problem = verify_photo_url("http://example.com/a.jpg", client)
    assert problem is not None
    assert "https" in problem


def test_verify_rejects_url_over_canva_char_limit() -> None:
    long_url = "https://example.com/" + "a" * MAX_IMAGE_URL_CHARS
    with _client(_ok) as client:
        problem = verify_photo_url(long_url, client)
    assert problem is not None
    assert str(MAX_IMAGE_URL_CHARS) in problem


def test_verify_rejects_403() -> None:
    """The exact Wikimedia failure that made the original defaults unusable."""

    def forbidden(_request: httpx.Request) -> httpx.Response:
        return httpx.Response(403, text="scripted request denied")

    with _client(forbidden) as client:
        problem = verify_photo_url("https://example.com/a.jpg", client)
    assert problem == "HTTP 403"


def test_verify_rejects_non_image_content_type() -> None:
    def html(_request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, headers={"content-type": "text/html"}, text="<html>")

    with _client(html) as client:
        problem = verify_photo_url("https://example.com/a.jpg", client)
    assert problem is not None
    assert "text/html" in problem


def test_verify_reports_transport_error_without_raising() -> None:
    def boom(_request: httpx.Request) -> httpx.Response:
        raise httpx.ConnectError("no route to host")

    with _client(boom) as client:
        problem = verify_photo_url("https://example.com/a.jpg", client)
    assert problem is not None
    assert "ConnectError" in problem


def test_csv_round_trips_with_commas_intact(tmp_path: Path) -> None:
    """Addresses and prices contain commas; quoting must survive a round trip."""
    path = tmp_path / "spike_a.csv"
    write_csv(path, build_rows(DEFAULT_PHOTO_URLS))
    with path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    assert tuple(rows[0]) == HEADERS
    assert rows[1][0] == ROW_TEXT[0][0]
    assert rows[1][1] == ROW_TEXT[0][1]


def test_xlsx_keeps_price_as_text(tmp_path: Path) -> None:
    """A price like $1,200,000 must not come back as a number -- why we prefer xlsx."""
    path = tmp_path / "spike_a.xlsx"
    write_xlsx(path, build_rows(DEFAULT_PHOTO_URLS))
    sheet = load_workbook(path).active
    assert sheet is not None
    values = list(sheet.iter_rows(values_only=True))
    assert tuple(values[0]) == HEADERS
    assert values[1][1] == "$1,200,000"
    assert isinstance(values[1][1], str)
    assert values[1][3] == DEFAULT_PHOTO_URLS[0]
