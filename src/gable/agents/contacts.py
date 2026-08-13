"""The agent roster, read from the working spreadsheet in the shared drive.

The roster used to be a `Sales_People` tab on the response workbook. It is now
`Agents Contact Information / Sales_Agents_Contact_Information.xlsx`, a document
Carmen and Chase edit by hand, which Gable reads on every pass and appends to
when it meets an agent nobody has recorded yet.

Two rules, both learned the hard way today:

* **The header is located, never assumed.** The old roster read from row 2, its
  header moved to row 1, and the sync silently stored nobody for a day — every
  flyer built that day carried the brokerage's main number and the design's own
  stock face instead of the agent's.
* **Nothing a human wrote is overwritten.** `append_contact` only ever adds a
  row. A phone number on a client-facing flyer is not something to correct
  automatically.

The workbook is parsed in memory. The droplet has 1 GB of RAM and the file is
about 25 KB, so a temporary file would buy nothing and leave litter.

Does not handle: two agents sharing an email. The first row wins and the
duplicate is logged, because picking between them is a filing question.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from sqlite3 import Connection
from typing import Any, Final

from gable.db import store
from gable.listings.headers import fold_header

logger = logging.getLogger("gable.agents.contacts")

#: The folder holding the roster, as a child of the configured Templates folder.
CONTACTS_FOLDER: Final[str] = "Agents Contact Information"

#: How far down to look for the header row before giving up.
MAX_HEADER_SCAN: Final[int] = 5

#: Folded header text to the field it fills. Only these four columns exist
#: today; anything else in the sheet is left alone.
HEADER_FIELDS: Final[dict[str, str]] = {
    "email": "email",
    "first name": "first_name",
    "last name": "last_name",
    "phone": "phone",
}


class ContactsError(Exception):
    """The roster workbook could not be read or written."""


@dataclass(frozen=True, slots=True)
class Contact:
    """One agent, as the roster records them."""

    email: str
    first_name: str
    last_name: str
    phone: str

    @property
    def full_name(self) -> str:
        """The name as a person writes it, for matching a headshot file."""
        return " ".join(part for part in (self.first_name, self.last_name) if part)


def find_workbook(drive: Any, drive_id: str, templates_folder_id: str) -> tuple[str, str]:  # noqa: ANN401
    """Locate the roster workbook inside the Templates folder.

    Args:
        drive: A Drive v3 resource.
        drive_id: The shared drive.
        templates_folder_id: The configured Templates folder.

    Returns:
        `(file_id, name)` of the first spreadsheet in the contacts folder.

    Raises:
        ContactsError: when the folder or the workbook is missing. A roster
            nobody can read is reported rather than treated as an empty one.
    """
    # Drive query grammar: https://developers.google.com/drive/api/guides/search-files
    folders = (
        drive.files()
        .list(
            corpora="drive",
            driveId=drive_id,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            q=(
                f"'{templates_folder_id}' in parents"
                " and mimeType='application/vnd.google-apps.folder'"
                f" and name='{CONTACTS_FOLDER}' and trashed=false"
            ),
            fields="files(id,name)",
            pageSize=10,
        )
        .execute()
        .get("files", [])
    )
    if not folders:
        msg = f"the {CONTACTS_FOLDER} folder is not in the drive"
        raise ContactsError(msg)
    files = (
        drive.files()
        .list(
            corpora="drive",
            driveId=drive_id,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            q=f"'{folders[0]['id']}' in parents and trashed=false",
            fields="files(id,name,mimeType)",
            pageSize=50,
        )
        .execute()
        .get("files", [])
    )
    books = [f for f in files if str(f.get("name", "")).lower().endswith(".xlsx")]
    if not books:
        msg = f"no contact workbook is in the {CONTACTS_FOLDER} folder"
        raise ContactsError(msg)
    return str(books[0]["id"]), str(books[0]["name"])


def _rows_of(data: bytes) -> list[list[str]]:
    """Read every populated cell of the first worksheet as text."""
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        msg = "the contact workbook could not be opened"
        raise ContactsError(msg) from exc
    try:
        sheet = workbook.worksheets[0]
        out: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            out.append(["" if cell is None else str(cell).strip() for cell in row])
        return out
    finally:
        workbook.close()


def find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Locate the header row and map its columns to contact fields.

    Args:
        rows: The worksheet's rows, from the first.

    Returns:
        `(index, columns)` — where the header sits, and field name to column.

    Raises:
        ContactsError: when no row in the first few names an email column.
    """
    for index, row in enumerate(rows[:MAX_HEADER_SCAN]):
        columns = {
            HEADER_FIELDS[folded]: position
            for position, cell in enumerate(row)
            if (folded := fold_header(cell)) in HEADER_FIELDS
        }
        if "email" in columns:
            return index, columns
    msg = "the contact workbook has no header row naming an email column"
    raise ContactsError(msg)


def parse_contacts(rows: list[list[str]]) -> list[Contact]:
    """Turn workbook rows into contacts, skipping anything without an email.

    Args:
        rows: The worksheet's rows.

    Returns:
        One contact per populated row, in sheet order, first row winning on a
        duplicated email.

    Raises:
        ContactsError: when the header cannot be found.
    """
    header_index, columns = find_header(rows)

    def cell(row: list[str], field: str) -> str:
        position = columns.get(field, -1)
        return row[position].strip() if 0 <= position < len(row) else ""

    out: list[Contact] = []
    seen: set[str] = set()
    for row in rows[header_index + 1 :]:
        email = cell(row, "email").lower()
        if not email:
            continue
        if email in seen:
            logger.warning("the contact workbook lists %s more than once; keeping the first", email)
            continue
        seen.add(email)
        out.append(
            Contact(
                email=email,
                first_name=cell(row, "first_name"),
                last_name=cell(row, "last_name"),
                phone=cell(row, "phone"),
            )
        )
    return out


def read_contacts(drive: Any, drive_id: str, templates_folder_id: str) -> list[Contact]:  # noqa: ANN401
    """Download and parse the roster workbook.

    Args:
        drive: A Drive v3 resource.
        drive_id: The shared drive.
        templates_folder_id: The configured Templates folder.

    Returns:
        Every agent the workbook records.

    Raises:
        ContactsError: when the workbook is missing or unreadable.
    """
    file_id, name = find_workbook(drive, drive_id, templates_folder_id)
    try:
        data = drive.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
    except Exception as exc:
        msg = f"{name} could not be downloaded"
        raise ContactsError(msg) from exc
    return parse_contacts(_rows_of(data))


def sync_contacts(
    drive: Any,  # noqa: ANN401 - googleapiclient resource, untyped upstream
    connection: Connection,
    drive_id: str,
    templates_folder_id: str,
) -> int:
    """Mirror the roster workbook into the local database.

    Args:
        drive: A Drive v3 resource.
        connection: An open database connection.
        drive_id: The shared drive.
        templates_folder_id: The configured Templates folder.

    Returns:
        How many agents were stored.

    Raises:
        ContactsError: when the workbook cannot be read. The caller decides
            whether that stops a run; it must never look like an empty roster.
    """
    contacts = read_contacts(drive, drive_id, templates_folder_id)
    for contact in contacts:
        store.upsert_salesperson(
            connection,
            email=contact.email,
            first_name=contact.first_name,
            last_name=contact.last_name,
            phone=contact.phone,
        )
    logger.info("mirrored %d agent(s) from the contact workbook", len(contacts))
    return len(contacts)


def append_contact(
    drive: Any,  # noqa: ANN401 - googleapiclient resource, untyped upstream
    drive_id: str,
    templates_folder_id: str,
    contact: Contact,
) -> bool:
    """Add one agent to the roster workbook, without touching any existing row.

    Args:
        drive: A Drive v3 resource.
        drive_id: The shared drive.
        templates_folder_id: The configured Templates folder.
        contact: The agent to record.

    Returns:
        True when the row was written, False when that email is already there.

    Raises:
        ContactsError: when the workbook cannot be read or written.
    """
    import openpyxl
    from googleapiclient.http import MediaIoBaseUpload

    file_id, name = find_workbook(drive, drive_id, templates_folder_id)
    try:
        data = drive.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
    except Exception as exc:
        msg = f"{name} could not be downloaded"
        raise ContactsError(msg) from exc

    rows = _rows_of(data)
    if any(c.email == contact.email.lower() for c in parse_contacts(rows)):
        return False
    _, columns = find_header(rows)

    # Written with the same workbook, not a fresh one, so anything else in the
    # file — other sheets, formatting, columns this code does not know about —
    # survives the round trip.
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data))
        sheet = workbook.worksheets[0]
        row = [""] * (max(columns.values()) + 1)
        for field, position in columns.items():
            row[position] = getattr(contact, field, "")
        sheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)
        drive.files().update(
            fileId=file_id,
            media_body=MediaIoBaseUpload(
                buffer,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=False,
            ),
            supportsAllDrives=True,
        ).execute()
    except Exception as exc:
        msg = f"{name} could not be updated"
        raise ContactsError(msg) from exc
    logger.info("added %s to the contact workbook", contact.email)
    return True
